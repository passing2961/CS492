import os
import numpy as np
import pickle as pc
from openpyxl import load_workbook

class EarlyStopping(object):
    """
    EarlyStopping
    """
    def __init__(self, best_loss, max_patience):
        super(EarlyStopping, self).__init__()
        
        self.best_loss = best_loss
        self.patience = 0
        self.max_patience = max_patience
        self.best_epoch = 0
        
    def update(self, recent_loss, recent_epoch):
        if self.best_loss > recent_loss:
            self.best_loss = recent_loss
            self.patience = 0
            
            self.best_epoch = recent_epoch
            return "update"
        else:
            self.patience += 1
            if self.patience == self.max_patience:
                return "terminate"
            
            return "patience"
        
    
class Average(object):
    """
    Keep track of most recent, average, sum, and count of a metric
    """
    def __init__(self):
        super(Average, self).__init__()
        self.reset()
        
    def reset(self):
        self.value = 0.
        self.avg = 0.
        self.sum = 0
        self.count = 0
    
    def update(self, value, n):
        self.value = value
        self.sum += value
        self.count = n
        self.avg = self.sum / self.count

def std_normalize(data):
    
    data = np.array(data)
    
    data_mean = data.mean()
    data_std = data.std()
    
    print("Mean: {}\tStd: {}".format(data_mean, data_std))
    return (data - data_mean) / data_std

def min_max(data):
    data = np.array(data)
    
    data_min = data.min()
    data_max = data.max()
    
    print("Min: {}\tMax: {}".format(data_min, data_max))
    return (data - data_min) / (data_max - data_min)

def missing_value(data):
    
    for i in range(len(data)):
        if data[i] > 100:
            data[i] = None
    
    for i in range(len(data)):
        if data[i] == None:
            if data[i-1] == None:
                prev = data[i-2]
                data[i] = (data[i-2] + data[i+1])/2
            elif data[i+1] == None:
                nxt = data[i+2]
                data[i] = (data[i-1] + data[i+2])/2
            else:
                data[i] = (data[i-1] + data[i+1])/2
                
    return data

def load_data(filename):
    '''
    Load xlsx data
    '''
    load_wb = load_workbook(filename, data_only=True)
    
    time = list()
    elec = list()
    # read elements per sheet
    #for ws in load_wb.worksheets:
    ws = load_wb['LBNL_B74_Electricity.csv']
    
    for idx, row in enumerate(ws.rows):
        if idx == 0:
            continue
            
        time.append(row[0].value)
        elec.append(row[1].value)
        
    print("Size of time, elec: {}\t{}".format(len(time), len(elec)))
    assert len(time) == len(elec)
    
    return time, elec

def load_gas_data(filename):
    load_wb = load_workbook(filename, data_only=True)
    
    gas = list()
    therms = list()
    time = list()
    
    ws = load_wb['LBNL B74 Gas Data']
    
    for idx, row in enumerate(ws.rows):
        if idx == 156629:
            break
            
        if idx == 0:
            continue
        
        time.append(row[0].value)
        gas.append(row[1].value)
        therms.append(row[2].value)
    print("Size of time, gas, therms: {}\t{}\t{}".format(len(time), len(gas), len(therms)))
    assert len(gas) == len(therms)
    return time, gas, therms

def split_dataset(data):
    train_data = data[:34967]
    test = data[34967:]
    
    idx = int(0.8 * len(train_data))
    
    train = train_data[:idx]
    valid = train_data[idx:]

    print("Size of train, valid, test: {}\t{}\t{}".format(len(train), len(valid), len(test)))
    return train, valid, test

def lstm_build_dataset(data, horizon_size):
    num_horizons = len(data) - horizon_size 
    
    observe = list()
    pred = list()
    for idx in range(num_horizons):
        start_idx = idx
        end_idx = idx + horizon_size
        
        observe.append(data[start_idx:end_idx])
        pred.append(data[start_idx+1:end_idx+1])
        #pred.append(data[end_idx+1])
        
    return observe, pred

def build_dataset(data, horizon_size):
    num_horizons = len(data) - horizon_size*2
    
    total_enc_data = list()
    total_dec_data = list()
    for idx in range(num_horizons):
        start_idx = idx
        mid_idx = start_idx + horizon_size
        end_idx = mid_idx + horizon_size
        
        enc_data = data[start_idx:mid_idx]
        dec_data = data[mid_idx:end_idx]
        #dec_tar_data = data[mid_idx+1:end_idx+1]
        total_enc_data.append(enc_data)
        total_dec_data.append(dec_data)
    
    total_enc_data = np.array(total_enc_data)
    total_dec_data = np.array(total_dec_data)
    
    assert len(total_enc_data) == len(total_dec_data)
    
    return total_enc_data, total_dec_data

def batch_iter(enc_data, dec_data, batch_size):
    num_batches_per_epoch = (len(enc_data) - 1) // batch_size + 1
    
    enc_data = np.array(enc_data)
    dec_data = np.array(dec_data)
    
    for batch_idx in range(num_batches_per_epoch):
        start_idx = batch_idx * batch_size
        end_idx = min((batch_idx + 1) * batch_size, len(enc_data))
        
        yield enc_data[start_idx:end_idx], dec_data[start_idx:end_idx]